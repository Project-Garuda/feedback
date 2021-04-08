from project.mod_faculty.models import Faculty,Courses , Feedback, Theory, Lab, Tutorial, UploadCourses, Admin
from project.mod_student.models import Student, UploadSection, Section
from project import db_session, bcrypt, app, Base, create_engine_models, delete_engine_models
from flask import Flask, render_template, request, redirect, url_for, Blueprint, session , abort, flash, g
from werkzeug.utils import secure_filename
import os
import openpyxl as op

mod_admin = Blueprint('admin', __name__)
@mod_admin.route("/dashboard" ,methods=['GET', 'POST'])
def admin_dashboard():
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        print(admin)
        upload_courses = UploadCourses.query.all()
        return render_template('admin/admin_dashboard.html',
        admin = admin,
        session = session,
        status = app.config['feedback_status'],
        upload_courses = upload_courses,
        faculty = Faculty,
        course_names = Courses
        )
    else:
        return redirect(url_for('admin_home'))

@mod_admin.route('/view/<int:id>', methods=['GET'])
def view_responses(id):
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        my_obj = UploadCourses.query.filter(UploadCourses.id == id).first()
        if my_obj.course == 0:
            theory = Theory.query.filter(Theory.id == id).first()
            if theory is None:
                abort(404)
            theory_dict = theory.fetch_dict()
            theory_dict['no_responses'] = theory.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(my_obj.id == Feedback.upload_courses_id).all()
            print(remarks)
            return render_template('admin/view_responses_theory.html',
            admin = admin,
            dict = theory_dict,
            remarks = remarks
            )
        elif my_obj.course == 1:
            lab = Lab.query.filter(Lab.id == id).first()
            print(lab)
            if lab is None:
                abort(404)
            lab_dict = lab.fetch_dict()
            lab_dict['no_responses'] = lab.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(my_obj.id == Feedback.upload_courses_id).all()
            print(remarks)
            return render_template('admin/view_responses_lab.html',
            admin = admin,
            dict = lab_dict,
            remarks = remarks
            )
        else:
            tutorial = Tutorial.query.filter(Tutorial.id == id).first()
            if tutorial is None:
                abort(404)
            tutorial_dict = tutorial.fetch_dict()
            tutorial_dict['no_responses'] = tutorial.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(my_obj.id == Feedback.upload_courses_id).all()
            print(remarks)
            return render_template('admin/view_responses_tutorial.html',
            admin = admin,
            dict = tutorial_dict,
            remarks = remarks
            )
    else:
        return redirect(url_for('admin_home'))

@mod_admin.route('/logout')
def logout():
    if 'admin' in session:
        session.pop('admin', None)
        flash('You have been logged out')
        return redirect(url_for('admin_home'))
    return redirect(url_for('admin_home'))

@mod_admin.route('/upload', methods=['GET', 'POST'])
def admin_upload():
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method == "POST":
            try:
                file_names = []
                for file in request.files:
                    current_file = request.files[file]
                    if current_file:
                        names=current_file.filename.split('.')
                        name=names[len(names)-1]
                        filename = secure_filename(file+'.'+name)
                        current_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        file_names.append(filename)
                        print('Uploaded', filename)
                if(process_data(file_names) == 1):
                    flash('Data processed Successfully')
                    return redirect(url_for('.admin_dashboard'))
                else:
                    flash('Error!')
                    return redirect(url_for('.admin_upload'))
            except Exception as e:
                print(e)
                flash('Unknown Error')
                return redirect(url_for('.admin_upload'))
        return render_template('admin/admin_upload.html', admin= admin)
    return redirect(url_for('admin_home'))

def process_data(file_names):
    admin_objs = Admin.query.all()
    try:
        delete_engine_models()
        create_engine_models()
        for obj in admin_objs:
            db_session.add(Admin(obj.id, obj.name, obj.password))
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error in clearing data')
        return redirect(url_for('.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], file_names[0]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = int(sheet_obj.cell(row = i, column = 1).value)
        name = sheet_obj.cell(row = i, column = 2).value
        password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
        db_session.add(Faculty(id, name, password))
    try:
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error in processing faculty data')
        return redirect(url_for('.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], file_names[1]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = int(sheet_obj.cell(row = i, column = 1).value)
        name = sheet_obj.cell(row = i, column = 2).value
        password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
        db_session.add(Student(id, name, password))
    try:
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error in processing student data')
        return redirect(url_for('.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], file_names[2]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = sheet_obj.cell(row = i, column = 1).value
        name = sheet_obj.cell(row = i, column = 2).value
        db_session.add(Courses(id, name))
    try:
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error in processing courses data')
        return redirect(url_for('.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], file_names[3]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column
    for i in range(1, m_col+1):
        id = sheet_obj.cell(row = 1, column = i).value
        db_session.add(Section(id))
        try:
            db_session.commit()
        except Exception as e:
            flash('Error in processing upload section data')
            return redirect(url_for('.admin_upload'))
        for j in range(2, m_row+1):
            obj = sheet_obj.cell(row = j, column = i).value
            if obj is None:
                break
            print(id, int(obj))
            db_session.add(UploadSection(id, int(obj)))
    try:
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error in processing upload section data')
        return redirect(url_for('.admin_upload'))

    return 1

@mod_admin.route('/add', methods=['GET', 'POST'])
def add_user():
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method == "POST":
            print('Hello')
            id = int(request.form['id'])
            name = request.form['name']
            role = request.form['role']
            password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
            if role == "Faculty":
                db_session.add(Faculty(id, name, password))
            if role == "Student":
                db_session.add(Student(id, name, password))
            try:
                db_session.commit()
                flash('Success')
                return redirect(url_for('.admin_dashboard'))
            except Exception as e:
                flash('Error! User already exists')
                return redirect(url_for('.add_user'))
        return render_template('admin/add_user.html',
        admin = admin,
        session = session,
        )
    else:
        return redirect(url_for('admin_home'))

@mod_admin.route('/change',methods=['GET', 'POST'])
def change_password():
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method == "POST":
            if bcrypt.check_password_hash(admin.password, request.form['old_pass']):
                new_pass = bcrypt.generate_password_hash(request.form['new_pass']).decode('utf-8')
                update_admin = Admin.query.filter(Admin.id == session['admin']).update({'password': new_pass})
                try:
                    db_session.commit()
                    flash('Successfully Updated Password')
                    return redirect(url_for('.admin_dashboard'))
                except Exception as e:
                    print(e)
                    flash('Unknown error!')
                    return redirect(url_for('.change_password'))
            else:
                flash('Old Password is incorrect')
                return redirect(url_for('.change_password'))
        return render_template('admin/change_password.html',
        admin = admin,
        session = session,
        )
    else:
        return redirect(url_for('admin_home'))

@mod_admin.route('/toggle', methods=['GET', 'POST'])
def toggle_feedback():
    if 'admin' in session:
        app.config['feedback_status'] = 1-app.config['feedback_status']
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        return redirect(url_for('.admin_dashboard'))
    else:
        return redirect(url_for('admin_home'))
